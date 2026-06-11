"""
Scrapling UI — Web Scraper Dashboard
Flask app with concurrent page fetching and smart duplicate detection.

Key improvements over v1:
  1. Pre-load DB fingerprints once per scrape run (set lookup, no per-job DB queries).
  2. Page-level duplicate tracking: stop only when N consecutive pages have ZERO new jobs.
     This correctly ignores sponsored/pinned jobs that repeat on every page.
  3. Concurrent page fetching via ThreadPoolExecutor (BATCH_SIZE=3 pages in parallel).
     Results are processed in page-order so stop logic remains accurate.
  4. Batch DB inserts per page — one transaction per page instead of one per job.
"""

import os
import sys
import json
import uuid
import threading
import traceback
import re
import time
import queue
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

from flask import Flask, render_template, request, jsonify, send_file
import database
import classifier
from dotenv import load_dotenv

load_dotenv()

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'scrapling-ui-secret-key-fallback')

database.init_db()

# In-memory job-run storage
scrape_jobs: dict = {}

# Batch queue processing
batch_queue = queue.Queue()

PORTAL_URL_CONFIGS = {
    'naukrigulf': {
        'UAE': 'https://www.naukrigulf.com/jobs-in-uae',
        'KSA': 'https://www.naukrigulf.com/jobs-in-saudi-arabia',
        'Qatar': 'https://www.naukrigulf.com/jobs-in-qatar',
        'Kuwait': 'https://www.naukrigulf.com/jobs-in-kuwait'
    },
    'linkedin': {
        'UAE': 'https://www.linkedin.com/jobs/search/?location=United%20Arab%20Emirates',
        'KSA': 'https://www.linkedin.com/jobs/search/?location=Saudi%20Arabia',
        'Qatar': 'https://www.linkedin.com/jobs/search/?location=Qatar',
        'Kuwait': 'https://www.linkedin.com/jobs/search/?location=Kuwait'
    },
    'gulftalent': {
        'UAE': 'https://www.gulftalent.com/uae/jobs',
        'KSA': 'https://www.gulftalent.com/saudi-arabia/jobs',
        'Qatar': 'https://www.gulftalent.com/qatar/jobs',
        'Kuwait': 'https://www.gulftalent.com/kuwait/jobs'
    },
    'bayt': {
        'UAE': 'https://www.bayt.com/en/uae/jobs/',
        'KSA': 'https://www.bayt.com/en/saudi-arabia/jobs/',
        'Qatar': 'https://www.bayt.com/en/qatar/jobs/',
        'Kuwait': 'https://www.bayt.com/en/kuwait/jobs/'
    }
}

def batch_worker():
    while True:
        task = batch_queue.get()
        if task is None:
            break
        
        master_id = task['master_id']
        site_type = task['site_type']
        country = task['country']
        url = task['url']
        max_pages = task['max_pages']
        
        master_job = scrape_jobs.get(master_id)
        if not master_job:
            batch_queue.task_done()
            continue
            
        if master_job.get('cancel_requested'):
            master_job['completed_tasks'] += 1
            if master_job['completed_tasks'] == master_job['total_tasks']:
                master_job['status'] = 'cancelled'
                master_job['completed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            batch_queue.task_done()
            continue
            
        master_job['message'] = f"Task {master_job['completed_tasks'] + 1}/{master_job['total_tasks']}: {country} on {site_type}..."
        
        try:
            dispatch = {
                'naukrigulf': scrape_naukrigulf,
                'linkedin':   scrape_linkedin,
                'gulftalent': scrape_gulftalent,
                'bayt':       scrape_bayt,
            }
            fn = dispatch.get(site_type, scrape_generic)
            
            results = fn(url, max_pages, country, master_id)
            
            if master_job.get('cancel_requested'):
                 master_job['status'] = 'cancelled'
            else:
                 master_job['results'].extend(results)
                 
        except Exception as e:
            print(f"[Batch Worker] Error on {country}/{site_type}: {e}")
            traceback.print_exc()
            
        master_job['completed_tasks'] += 1
        
        if master_job.get('cancel_requested'):
            master_job['status'] = 'cancelled'
            master_job['completed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elif master_job['completed_tasks'] == master_job['total_tasks']:
            master_job['status'] = 'completed'
            master_job['message'] = f"Batch complete! Found {len(master_job['results'])} total items."
            master_job['completed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            if not master_job.get('cancel_requested'):
                master_job['message'] = f"Waiting 5s before next task..."
                time.sleep(5)
            
        batch_queue.task_done()

threading.Thread(target=batch_worker, daemon=True).start()

# ─── Constants ────────────────────────────────────────────────────────────────

# How many pages to fetch in parallel.
# 3 is safe — aggressive enough to be fast, gentle enough to avoid blocks.
CONCURRENT_PAGES = 3

# Stop when this many consecutive pages contain ZERO new jobs.
# 2 means: we need 2 full pages of 100% old data before we declare "done".
# This handles transition pages (e.g. page where 9 are old, 1 is new).
STOP_AFTER_DUPE_PAGES = 2


# ─── Utilities ────────────────────────────────────────────────────────────────

def get_absolute_date(relative_str: str) -> str:
    """Convert relative time strings to YYYY-MM-DD format."""
    if not relative_str:
        return ""

    relative_str = relative_str.lower().strip()
    now = datetime.now()

    if any(x in relative_str for x in ['active', 'just now', 'today']):
        return now.strftime('%Y-%m-%d')

    if 'yesterday' in relative_str:
        return (now - timedelta(days=1)).strftime('%Y-%m-%d')

    match = re.search(r'(\d+)\s*(min|hour|hr|day|week|month|year)s?', relative_str)
    if match:
        val = int(match.group(1))
        unit = match.group(2)
        if unit == 'min':
            return now.strftime('%Y-%m-%d')
        elif unit in ['hour', 'hr']:
            return now.strftime('%Y-%m-%d')
        elif unit == 'day':
            return (now - timedelta(days=val)).strftime('%Y-%m-%d')
        elif unit == 'week':
            return (now - timedelta(weeks=val)).strftime('%Y-%m-%d')
        elif unit == 'month':
            return (now - timedelta(days=val * 30)).strftime('%Y-%m-%d')
        elif unit == 'year':
            return (now - timedelta(days=val * 365)).strftime('%Y-%m-%d')

    try:
        clean_str = relative_str
        for fmt in ('%d %b %Y', '%d %B %Y', '%d %b', '%d %B',
                    '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%d'):
            try:
                dt = datetime.strptime(clean_str, fmt)
                if '%Y' not in fmt:
                    dt = dt.replace(year=now.year)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
    except Exception:
        pass

    return relative_str


def make_fingerprint(title: str, company: str, location: str, source: str, country: str = '') -> str:
    """Must match database._make_fingerprint exactly."""
    return (
        f"{(title or '').lower().strip()}"
        f"|{(company or '').lower().strip()}"
        f"|{(location or '').lower().strip()}"
        f"|{(source or '').lower().strip()}"
        f"|{(country or '').lower().strip()}"
    )


def parse_relative_time(time_str: str) -> float:
    """Convert time string to minutes-ago float (for sorting)."""
    if not time_str:
        return float('inf')

    time_str = str(time_str).lower()
    if 'active' in time_str or 'just now' in time_str:
        return 0

    match = re.search(r'(\d+)\s*(min|hour|hr|day|week|month|year)s?', time_str)
    if match:
        val = int(match.group(1))
        unit = match.group(2)
        if unit == 'min':      return val
        if unit in ['hour', 'hr']: return val * 60
        if unit == 'day':      return val * 24 * 60
        if unit == 'week':     return val * 7 * 24 * 60
        if unit == 'month':    return val * 30 * 24 * 60
        if unit == 'year':     return val * 365 * 24 * 60

    try:
        dt = datetime.strptime(time_str, '%Y-%m-%d')
        return (datetime.now() - dt).total_seconds() / 60
    except Exception:
        pass

    return float('inf')


# ─── Low-level page fetcher (runs in thread pool) ─────────────────────────────

def _fetch_page(url: str, page_num: int, site: str = ''):
    from scrapling.fetchers import StealthyFetcher
    try:
        # Adding a 1-2 second artificial delay to slow down overall scraping
        time.sleep(1.5)
        
        # Naukrigulf is a heavy SPA and requires waiting for network requests to settle 
        # so the JS can render the job cards.
        wait_network = True if site == 'naukrigulf' else False
        
        print(f"  → Fetching page {page_num}: {url}")
        page = StealthyFetcher.fetch(
            url,
            headless=True,
            network_idle=wait_network,
            block_images=True,
            timeout=45000,
        )
        return page_num, page
    except Exception as e:
        print(f"  ✗ Page {page_num} failed: {e}")
        return page_num, None


def _fetch_pages_concurrent(page_specs: list[tuple[int, str]], site: str = '') -> dict:
    """
    Fetch a batch of (page_num, url) pairs concurrently.
    Returns dict {page_num: page_object_or_None}.
    """
    results = {}
    with ThreadPoolExecutor(max_workers=len(page_specs)) as executor:
        futures = {}
        for idx, (pn, url) in enumerate(page_specs):
            if idx > 0:
                time.sleep(2.0)  # Stagger concurrent requests to avoid rate-limiting
            future = executor.submit(_fetch_page, url, pn, site)
            futures[future] = pn
            
        for future in as_completed(futures):
            try:
                pn, page = future.result()
                results[pn] = page
            except Exception as e:
                print(f"  ✗ Thread error: {e}")
    return results


# ─── Per-page job processors ──────────────────────────────────────────────────
# Each returns a list of job dicts extracted from the page HTML.

def _parse_naukrigulf_page(page, page_num: int) -> list[dict]:
    jobs = []
    # Broaden card selectors to catch various structures used by Naukrigulf
    cards = page.css('.srp-tuple, .tuple-wrap, .scroll-tuple, [class*="tuple"], .srp-job-tuple')
    
    for card in cards:
        title_el = (
            card.css('.designation-title::text') or 
            card.css('.title::text') or
            card.css('[class*="title"]::text') or
            card.css('h2::text')
        )
        
        company_el = (
            card.css('.info-org::text') or 
            card.css('.company-name::text') or 
            card.css('[class*="company"]::text') or
            card.css('p.m-0::text')
        )
        
        spans = card.css('span:not([class])::text').getall()
        loc_el = card.css('.info-loc::text') or card.css('[class*="loc"]::text')
        exp_el = card.css('.info-exp::text') or card.css('[class*="exp"]::text')
        
        time_el = card.css('.time::text') or card.css('.date::text') or card.css('[class*="time"]::text')

        title_text = (title_el.get('') if title_el else '').strip()
        if not title_text:
            links = card.css('a::text').getall()
            for l in links:
                if len(l.strip()) > 5:
                    title_text = l.strip()
                    break

        if not title_text:
            continue

        company_text = (company_el.get('') if company_el else '').strip()
        
        loc_text = (loc_el.get('') if loc_el else '').strip()
        if not loc_text and len(spans) > 1:
            loc_text = spans[1].strip()
            
        exp_text = (exp_el.get('') if exp_el else '').strip()
        if not exp_text and len(spans) > 0:
            exp_text = spans[0].strip()

        job = {
            'title':      title_text,
            'company':    company_text,
            'location':   loc_text,
            'experience': exp_text,
            'posted':     get_absolute_date((time_el.get('') if time_el else '').strip()),
            'source':     'Naukrigulf',
            'page':       page_num,
        }
        jobs.append(job)
    return jobs


def _parse_linkedin_page(page, page_num: int) -> list[dict]:
    jobs = []
    cards = (
        page.css('.base-card') or
        page.css('.job-search-card') or
        page.css('[data-entity-urn*="jobPosting"]') or
        page.css('.jobs-search__results-list li') or
        page.css('.result-card') or
        page.css('ul.jobs-search__results-list > li')
    )
    for card in cards:
        title_el = (
            card.css('.base-search-card__title::text') or
            card.css('.job-search-card__title::text') or
            card.css('h3::text') or
            card.css('[class*="title"]::text')
        )
        company_el = (
            card.css('.base-search-card__subtitle a::text') or
            card.css('.base-search-card__subtitle::text') or
            card.css('a.hidden-nested-link::text') or
            card.css('.job-search-card__subtitle-link::text') or
            card.css('.base-search-card__subtitle [data-tracking-control-name*="subtitle"]::text') or
            card.css('h4 a::text') or
            card.css('h4::text') or
            card.css('[class*="company"]::text')
        )
        loc_el = (
            card.css('.job-search-card__location::text') or
            card.css('.base-search-card__metadata span::text') or
            card.css('[class*="location"]::text')
        )
        date_el    = (
            card.css('time::text') or
            card.css('.job-search-card__listdate::text') or
            card.css('[class*="date"]::text')
        )
        posted_raw = (date_el.get('') if date_el else '').strip()
        if not posted_raw:
            time_tag = card.css('time')
            if time_tag:
                posted_raw = time_tag[0].attrib.get('datetime', '')

        job = {
            'title':      (title_el.get('') if title_el else '').strip(),
            'company':    (company_el.get('') if company_el else '').strip(),
            'location':   (loc_el.get('') if loc_el else '').strip(),
            'experience': '',
            'posted':     get_absolute_date(posted_raw),
            'source':     'LinkedIn',
            'page':       page_num,
        }
        if job['title']:
            jobs.append(job)
    return jobs


def _parse_gulftalent_page(page, page_num: int) -> list[dict]:
    jobs = []
    # GulfTalent uses table rows for job results
    rows = page.css('tr.content-visibility-auto') or page.css('.job-result-row') or page.css('table tr')
    for row in rows:
        title_el = (
            row.css('h2.title a::text') or 
            row.css('.job-title a::text') or 
            row.css('a.title::text') or 
            row.css('a.text-title::text')
        )
        title = (title_el.get('') if title_el else '').strip()
        if not title:
            continue
        
        # Company is usually a link with class text-muted or contains /companies/ in href
        company_el = (
            row.css('a.text-muted::text') or 
            row.css('a[href*="/companies/"]::text') or
            row.css('.company-name::text')
        )
        company = (company_el.get('') if company_el else '').strip()
        
        # Robust fallback: check all links in the row and take the first one that isn't the title
        if not company:
            links = row.css('a::text').getall()
            for l in links:
                l = l.strip()
                if l and l != title and len(l) > 1:
                    company = l
                    break

        loc_el = (
            row.css('td.col-sm-6 span::text') or 
            row.css('.location::text') or 
            row.css('a.text-regular::text') or
            row.css('td:nth-child(2) ::text')
        )
        date_el = (
            row.css('td.col-sm-4::text') or 
            row.css('.date-posted::text') or
            row.css('td:last-child ::text')
        )

        job = {
            'title':      title,
            'company':    company,
            'location':   (loc_el.get('') if loc_el else '').strip(),
            'experience': '',
            'posted':     get_absolute_date((date_el.get('') if date_el else '').strip()),
            'source':     'GulfTalent',
            'page':       page_num,
        }
        if job['title']:
            jobs.append(job)
    return jobs


def _parse_bayt_page(page, page_num: int) -> list[dict]:
    jobs = []
    cards = page.css('li.has-pointer-d')
    for card in cards:
        title_el   = card.css('h2 a::text')
        company_el = card.css('.job-company-location-wrapper a.t-bold::text')
        loc_el     = card.css('.job-company-location-wrapper div.t-small a:first-child span::text')
        date_el    = card.css('.jb-date span::text')
        exp_el     = card.css('.jb-label-careerlevel::text')

        title = (title_el.get('') if title_el else '').strip()
        if not title or 'Mobile App' in title:
            continue

        job = {
            'title':      title,
            'company':    (company_el.get('') if company_el else '').strip(),
            'location':   (loc_el.get('') if loc_el else '').strip(),
            'experience': (exp_el.get('') if exp_el else '').strip(),
            'posted':     get_absolute_date((date_el.get('') if date_el else '').strip()),
            'source':     'Bayt',
            'page':       page_num,
        }
        jobs.append(job)
    return jobs


# ─── Core scraping engine ─────────────────────────────────────────────────────

def _build_page_url(base_url: str, page_num: int, site: str) -> str:
    """Return the correct paginated URL for a given site and page number."""
    if site == 'naukrigulf':
        if page_num == 1:
            return base_url
        if '?' in base_url:
            return f"{base_url}&pageNo={page_num}"
        return f"{base_url.rstrip('/')}-{page_num}"

    elif site == 'linkedin':
        start = (page_num - 1) * 25
        if 'linkedin.com/jobs/search' in base_url:
            api_base = base_url.replace(
                'linkedin.com/jobs/search',
                'linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search'
            )
            sep = '&' if '?' in api_base else '?'
            return f"{api_base}{sep}start={start}"
        if page_num == 1:
            return base_url
        sep = '&' if '?' in base_url else '?'
        return f"{base_url}{sep}start={start}"

    elif site == 'gulftalent':
        if page_num == 1:
            return base_url
        return f"{base_url.rstrip('/')}/{page_num}"

    elif site == 'bayt':
        if page_num == 1:
            return base_url
        sep = '&' if '?' in base_url else '?'
        return f"{base_url}{sep}page={page_num}"

    return base_url


def _get_parser(site: str):
    parsers = {
        'naukrigulf': _parse_naukrigulf_page,
        'linkedin':   _parse_linkedin_page,
        'gulftalent': _parse_gulftalent_page,
        'bayt':       _parse_bayt_page,
    }
    return parsers.get(site)


def _scrape_site(url: str, site: str, max_pages: int, country: str = '', job_id: str = None) -> list[dict]:
    """
    Generic concurrent scraping engine used by all site scrapers.
    country — e.g. 'UAE', 'KSA', 'Qatar', 'Kuwait' — stored in DB for per-country filtering.
    job_id — used for progress reporting and cancellation checks.
    """
    source_name = {
        'naukrigulf': 'Naukrigulf',
        'linkedin':   'LinkedIn',
        'gulftalent': 'GulfTalent',
        'bayt':       'Bayt',
    }.get(site, site.title())

    parser = _get_parser(site)
    if not parser:
        print(f"[!] No parser for site: {site}")
        return []

    # Step 1: Pre-load fingerprints scoped to this source+country
    db_fingerprints: set = database.load_fingerprints(source_name, country or None)

    all_new_jobs: list[dict] = []
    seen_in_run: set         = set()
    consecutive_dupe_pages   = 0
    page_num                 = 1

    while page_num <= max_pages:
        batch: list[tuple[int, str]] = []
        for i in range(CONCURRENT_PAGES):
            pn = page_num + i
            if pn > max_pages:
                break
            batch.append((pn, _build_page_url(url, pn, site)))

        if not batch:
            break

        print(f"\n[{source_name}/{country or 'all'}] Fetching pages {batch[0][0]}–{batch[-1][0]}...")

        fetched = _fetch_pages_concurrent(batch, site)

        should_stop = False
        for pn, page_url in batch:
            # Check for cancellation
            if scrape_jobs.get(job_id, {}).get('cancel_requested'):
                print(f"  ⚠ Cancellation requested for job {job_id} at page {pn}.")
                scrape_jobs[job_id]['status'] = 'cancelled'
                scrape_jobs[job_id]['message'] = 'Scraping cancelled by user'
                should_stop = True
                break

            if scrape_jobs.get(job_id, {}).get('is_batch'):
                task_msg = f"[{country}/{site}] "
                if max_pages >= 9000:
                    scrape_jobs[job_id]['message'] = f'{task_msg}Processing page {pn}...'
                else:
                    scrape_jobs[job_id]['message'] = f'{task_msg}Processing page {pn} of {max_pages}...'
            else:
                if max_pages >= 9000:
                    scrape_jobs[job_id]['message'] = f'Processing page {pn}...'
                else:
                    scrape_jobs[job_id]['message'] = f'Processing page {pn} of {max_pages}...'
                
            page = fetched.get(pn)
            if page is None:
                print(f"  ✗ Page {pn} returned None — stopping.")
                should_stop = True
                break

            raw_jobs = parser(page, pn)
            if not raw_jobs:
                print(f"  ✗ Page {pn} had no job cards — stopping.")
                should_stop = True
                break

            new_this_page: list[dict] = []
            new_count  = 0
            dupe_count = 0

            for job in raw_jobs:
                session_key = (
                    job['title'].lower().strip(),
                    job['company'].lower().strip(),
                    job['location'].lower().strip(),
                )
                if session_key in seen_in_run:
                    continue
                seen_in_run.add(session_key)

                # Stamp country and scrape_id onto the job dict
                job['country'] = database.normalize_country(country)
                job['scrape_id'] = job_id

                fp = make_fingerprint(
                    job['title'], job['company'], job['location'], job['source'], country
                )

                # ICT Classification & Date Normalization
                is_ict, ict_cat = classifier.classify_job(job['title'])
                norm_date, precision = normalize_posted_date(job.get('posted', ''))
                
                job['_is_ict'] = is_ict
                job['ict_category'] = ict_cat
                job['posted'] = norm_date
                job['date_precision'] = precision

                if fp in db_fingerprints:
                    dupe_count += 1
                else:
                    new_count += 1
                    new_this_page.append(job)
                    db_fingerprints.add(fp)

            print(
                f"  ✓ Page {pn}: {new_count} new | {dupe_count} already-in-db "
                f"| dupe-page streak: {consecutive_dupe_pages}"
            )

            if new_this_page:
                database.save_jobs_batch(new_this_page)
                all_new_jobs.extend(new_this_page)
                consecutive_dupe_pages = 0
            else:
                if dupe_count > 0:
                    consecutive_dupe_pages += 1
                    print(
                        f"  ⚠ Page {pn} is 100% old data. "
                        f"Streak: {consecutive_dupe_pages}/{STOP_AFTER_DUPE_PAGES}"
                    )
                    if consecutive_dupe_pages >= STOP_AFTER_DUPE_PAGES:
                        print(f"  ⛔ {STOP_AFTER_DUPE_PAGES} consecutive pages of old data — stopping.")
                        should_stop = True
                        break
                else:
                    # 0 new AND 0 old unique jobs — usually means end of results or empty redirect
                    # User requested: "instead of stopping by saying 0 jobs"
                    # We will be slightly more patient here.
                    consecutive_dupe_pages += 1
                    print(f"  ⚠ Page {pn} has 0 unique jobs. Streak: {consecutive_dupe_pages}/{STOP_AFTER_DUPE_PAGES}")
                    if consecutive_dupe_pages >= STOP_AFTER_DUPE_PAGES:
                        print(f"  ⛔ {STOP_AFTER_DUPE_PAGES} consecutive empty/dupe pages — stopping.")
                        should_stop = True
                        break

        if should_stop:
            break

        page_num += CONCURRENT_PAGES

    print(f"\n[{source_name}/{country or 'all'}] Done. {len(all_new_jobs)} new jobs found.")
    return all_new_jobs


# ─── Public scraper functions ─────────────────────────────────────────────────

def scrape_naukrigulf(url: str, max_pages: int = 5, country: str = '', job_id: str = None) -> list[dict]:
    return _scrape_site(url, 'naukrigulf', max_pages, country, job_id)


def scrape_linkedin(url: str, max_pages: int = 5, country: str = '', job_id: str = None) -> list[dict]:
    return _scrape_site(url, 'linkedin', max_pages, country, job_id)


def scrape_gulftalent(url: str, max_pages: int = 5, country: str = '', job_id: str = None) -> list[dict]:
    return _scrape_site(url, 'gulftalent', max_pages, country, job_id)


def scrape_bayt(url: str, max_pages: int = 5, country: str = '', job_id: str = None) -> list[dict]:
    return _scrape_site(url, 'bayt', max_pages, country, job_id)


def scrape_generic(url: str) -> list[dict]:
    """Scrape a generic URL and extract structured data."""
    from scrapling.fetchers import Fetcher

    all_data = []
    try:
        page = Fetcher.get(url, stealthy_headers=True, follow_redirects=True)
        if page is None:
            return all_data

        # Try tables first
        tables = page.css('table')
        if tables:
            for table in tables:
                headers = [th.css('::text').get('').strip() for th in table.css('th')]
                for row in table.css('tr'):
                    cells = row.css('td')
                    if cells:
                        row_data = {}
                        for i, cell in enumerate(cells):
                            key = headers[i] if i < len(headers) else f'Column {i + 1}'
                            row_data[key] = cell.css('::text').get('').strip()
                        if any(row_data.values()):
                            all_data.append(row_data)
            return all_data

        # Fall back to card/item heuristics
        items = (
            page.css('article') or
            page.css('[class*="card"]') or
            page.css('[class*="item"]') or
            page.css('[class*="result"]') or
            page.css('li')
        )

        for item in items[:50]:
            data = {}
            heading = item.css('h1::text, h2::text, h3::text, h4::text').get('')
            if heading:
                data['title'] = heading.strip()

            link = item.css('a')
            if link:
                data['link'] = link[0].attrib.get('href', '')
                if not data.get('title'):
                    data['title'] = link[0].css('::text').get('').strip()

            para = item.css('p::text').get('')
            if para:
                data['description'] = para.strip()

            for i, span in enumerate(item.css('span::text').getall()[:5]):
                if span.strip():
                    data[f'detail_{i + 1}'] = span.strip()

            data['source'] = 'Custom'

            if data.get('title'):
                ordered = {
                    'title':      data.get('title', ''),
                    'company':    data.get('company', ''),
                    'location':   data.get('location', ''),
                    'experience': data.get('experience', ''),
                    'posted':     get_absolute_date(data.get('posted', '')),
                    'source':     'Custom',
                }
                if database.save_job(ordered):
                    all_data.append(ordered)

    except Exception as e:
        print(f"[Generic] Error: {e}")
        traceback.print_exc()

    return all_data


# ─── Site detection ───────────────────────────────────────────────────────────

def detect_site(url: str) -> str:
    url_lower = url.lower()
    if 'naukrigulf' in url_lower: return 'naukrigulf'
    if 'linkedin'   in url_lower: return 'linkedin'
    if 'gulftalent' in url_lower: return 'gulftalent'
    if 'bayt'       in url_lower: return 'bayt'
    return 'generic'


# ─── Background runner ────────────────────────────────────────────────────────

def run_scrape(job_id: str, url: str, max_pages: int, site_type: str, country: str = ''):
    """Run scraping in a background thread and update scrape_jobs[job_id]."""
    try:
        scrape_jobs[job_id]['status']  = 'running'
        scrape_jobs[job_id]['message'] = f'Starting {site_type} scrape...'

        dispatch = {
            'naukrigulf': scrape_naukrigulf,
            'linkedin':   scrape_linkedin,
            'gulftalent': scrape_gulftalent,
            'bayt':       scrape_bayt,
        }
        fn = dispatch.get(site_type, scrape_generic)

        if site_type == 'generic':
            results = fn(url)
        else:
            results = fn(url, max_pages, country, job_id)

        scrape_jobs[job_id]['results']      = results
        scrape_jobs[job_id]['status']       = 'completed'
        scrape_jobs[job_id]['message']      = f'Found {len(results)} new items'
        scrape_jobs[job_id]['completed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    except Exception as e:
        scrape_jobs[job_id]['status']  = 'error'
        scrape_jobs[job_id]['message'] = str(e)
        traceback.print_exc()


# ─── Excel export ─────────────────────────────────────────────────────────────

def _excel_styles():
    return {
        'header_font': Font(name='Segoe UI', bold=True, color='FFFFFF', size=11),
        'header_fill': PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid'),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'cell_font': Font(name='Segoe UI', size=10),
        'cell_alignment': Alignment(vertical='top', wrap_text=True),
        'thin_border': Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0'),
        ),
        'alt_fill': PatternFill(start_color='F5F5FA', end_color='F5F5FA', fill_type='solid'),
    }


def _is_ict_job(row: dict) -> bool:
    val = row.get('is_ict', row.get('_is_ict', 0))
    return val in (True, 1, '1', 'True')


def _write_data_sheet(ws, group_rows: list[dict], column_order: list[str], skip_keys: set, styles: dict):
    """Populate one worksheet with headers, rows, column widths."""
    actual_keys = [k for k in column_order if any(k in row for row in group_rows)]
    for row in group_rows:
        for k in row.keys():
            if k not in actual_keys and k not in skip_keys:
                actual_keys.append(k)

    for col, key in enumerate(actual_keys, 1):
        cell = ws.cell(row=1, column=col, value=key.replace('_', ' ').title())
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['thin_border']

    for row_idx, row_data in enumerate(group_rows, 2):
        for col_idx, key in enumerate(actual_keys, 1):
            val = row_data.get(key, '')
            if key == 'is_ict' and val == '':
                val = row_data.get('_is_ict', '')
            if key == 'is_ict':
                val = 'Yes' if val in (True, 1, '1', 'True') else 'No'
            if key == 'country' and val:
                val = database.normalize_country(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = styles['cell_font']
            cell.alignment = styles['cell_alignment']
            cell.border = styles['thin_border']
            if row_idx % 2 == 0:
                cell.fill = styles['alt_fill']

    for col_idx, key in enumerate(actual_keys, 1):
        max_length = len(key) + 4
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

    ws.freeze_panes = 'A2'


def generate_excel(data: list[dict], filename: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    os.makedirs(os.path.join(os.path.dirname(__file__), 'exports'), exist_ok=True)
    filepath = os.path.join(os.path.dirname(__file__), 'exports', filename)

    if not data:
        ws['A1'] = 'No data found'
        wb.save(filepath)
        return filepath

    column_order = ['title', 'company', 'location', 'experience', 'posted', 'source', 'is_ict', 'ict_category']
    actual_keys  = [k for k in column_order if any(k in row for row in data)]

    header_font      = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    header_fill      = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font        = Font(name='Segoe UI', size=10)
    cell_alignment   = Alignment(vertical='top', wrap_text=True)
    thin_border      = Border(
        left   = Side(style='thin', color='E0E0E0'),
        right  = Side(style='thin', color='E0E0E0'),
        top    = Side(style='thin', color='E0E0E0'),
        bottom = Side(style='thin', color='E0E0E0'),
    )
    alt_fill = PatternFill(start_color='F5F5FA', end_color='F5F5FA', fill_type='solid')

    for col, key in enumerate(actual_keys, 1):
        cell = ws.cell(row=1, column=col, value=key.replace('_', ' ').title())
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_alignment
        cell.border    = thin_border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(actual_keys, 1):
            val = row_data.get(key, '')
            # Handle key variations (_is_ict vs is_ict)
            if key == 'is_ict' and val == '':
                val = row_data.get('_is_ict', '')
            
            # Convert boolean/int to Yes/No for better readability in Excel
            if key == 'is_ict':
                val = "Yes" if val in [True, 1, "1", "True"] else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = cell_font
            cell.alignment = cell_alignment
            cell.border    = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, key in enumerate(actual_keys, 1):
        max_length = len(key) + 4
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

    ws.freeze_panes = 'A2'
    wb.save(filepath)
    return filepath

def generate_multi_excel(data: list[dict], filename: str) -> str:
    from collections import defaultdict
    wb = Workbook()
    
    os.makedirs(os.path.join(os.path.dirname(__file__), 'exports'), exist_ok=True)
    filepath = os.path.join(os.path.dirname(__file__), 'exports', filename)

    if not data:
        ws = wb.active
        ws.title = "Scraped Data"
        ws['A1'] = 'No data found'
        wb.save(filepath)
        return filepath

    # Group data by (source, country)
    grouped_data = defaultdict(list)
    for row in data:
        source = row.get('source', 'Unknown').capitalize()
        country = database.normalize_country(row.get('country', '')) or 'General'
        grouped_data[(source, country)].append(row)

    styles = _excel_styles()
    portal_column_order = ['title', 'company', 'location', 'experience', 'posted', 'is_ict', 'ict_category']
    ict_column_order = ['title', 'company', 'location', 'experience', 'posted', 'source', 'country', 'is_ict', 'ict_category']
    portal_skip = {'source', 'country', 'scrape_id', 'id', 'scraped_at'}

    default_sheet = wb.active
    wb.remove(default_sheet)

    ict_rows = [r for r in data if _is_ict_job(r)]
    if ict_rows:
        ws_ict = wb.create_sheet(title='All ICT Jobs', index=0)
        _write_data_sheet(ws_ict, ict_rows, ict_column_order, {'scrape_id', 'id'}, styles)

    for (source, country), group_rows in grouped_data.items():
        sheet_title = f"{source[:15]} - {country[:12]}"
        ws = wb.create_sheet(title=sheet_title)
        _write_data_sheet(ws, group_rows, portal_column_order, portal_skip, styles)

    wb.save(filepath)
    return filepath


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/scrape', methods=['POST'])
def start_scrape():
    data      = request.json
    url       = data.get('url', '').strip()
    max_pages = int(data.get('max_pages', 5))
    country   = data.get('country', '').strip()   # e.g. 'UAE', 'KSA', 'Qatar', 'Kuwait'

    if not url:
        return jsonify({'error': 'URL is required'}), 400

    site_type = data.get('site_type', 'auto')
    if site_type == 'auto':
        site_type = detect_site(url)

    job_id = str(uuid.uuid4())[:8]
    scrape_jobs[job_id] = {
        'id':           job_id,
        'url':          url,
        'site_type':    site_type,
        'country':      country,
        'status':       'queued',
        'message':      'Starting...',
        'results':      [],
        'created_at':   datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'completed_at': None,
        'max_pages':    max_pages,
    }

    thread = threading.Thread(target=run_scrape, args=(job_id, url, max_pages, site_type, country))
    thread.daemon = True
    thread.start()

    return jsonify({'job_id': job_id, 'site_type': site_type, 'country': country})


@app.route('/api/scrape/batch', methods=['POST'])
def start_batch_scrape():
    data = request.json
    countries = data.get('countries', [])
    portals = data.get('portals', [])
    max_pages = int(data.get('max_pages', 5))

    if not countries or not portals:
        return jsonify({'error': 'Countries and Portals are required'}), 400

    job_id = str(uuid.uuid4())[:8]
    total_tasks = len(countries) * len(portals)

    scrape_jobs[job_id] = {
        'id':           job_id,
        'url':          'Batch Job',
        'site_type':    'batch',
        'country':      'Multiple',
        'status':       'running',
        'message':      f'Queued {total_tasks} tasks...',
        'results':      [],
        'created_at':   datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'completed_at': None,
        'max_pages':    max_pages,
        'is_batch':     True,
        'total_tasks':  total_tasks,
        'completed_tasks': 0,
    }

    for country in countries:
        for portal in portals:
            # Resolve URL from config
            url = PORTAL_URL_CONFIGS.get(portal, {}).get(country)
            if not url:
                print(f"[Batch] WARNING: No URL config for {portal}/{country}")
                # We still queue it so the total tasks count doesn't mismatch, 
                # but it will likely fail gracefully or we can just decrement total_tasks.
                # It's better to push a task and let the worker handle it or skip.
                scrape_jobs[job_id]['total_tasks'] -= 1
                continue
                
            task = {
                'master_id': job_id,
                'site_type': portal,
                'country': country,
                'url': url,
                'max_pages': max_pages
            }
            batch_queue.put(task)

    return jsonify({'job_id': job_id, 'site_type': 'batch', 'country': 'Multiple'})


@app.route('/api/cancel/<job_id>', methods=['POST'])
def cancel_scrape(job_id):
    if job_id in scrape_jobs:
        scrape_jobs[job_id]['cancel_requested'] = True
        return jsonify({'status': 'cancel_requested'})
    return jsonify({'error': 'Job not found'}), 404


@app.route('/api/status/<job_id>')
def get_status(job_id):
    job = scrape_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404

    return jsonify({
        'id':           job['id'],
        'status':       job['status'],
        'message':      job['message'],
        'result_count': len(job['results']),
        'results':      job['results'],
        'site_type':    job['site_type'],
        'url':          job['url'],
        'created_at':   job['created_at'],
        'completed_at': job['completed_at'],
    })


@app.route('/api/export/<job_id>')
def export_excel(job_id):
    job = scrape_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    if not job['results']:
        return jsonify({'error': 'No data to export'}), 400

    results_to_export = list(job['results'])
    sort_mode = request.args.get('sort', 'default')

    if sort_mode == 'recent':
        results_to_export.sort(key=lambda x: parse_relative_time(x.get('posted', '')))
    elif sort_mode == 'oldest':
        results_to_export.sort(key=lambda x: parse_relative_time(x.get('posted', '')), reverse=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f"scraped_{job['site_type']}_{timestamp}.xlsx"
    filepath  = generate_excel(results_to_export, filename)
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/api/export_multi/<job_id>')
def export_multi_excel(job_id):
    job = scrape_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    if not job['results']:
        return jsonify({'error': 'No data to export'}), 400

    results_to_export = list(job['results'])
    sort_mode = request.args.get('sort', 'default')

    if sort_mode == 'recent':
        results_to_export.sort(key=lambda x: parse_relative_time(x.get('posted', '')))
    elif sort_mode == 'oldest':
        results_to_export.sort(key=lambda x: parse_relative_time(x.get('posted', '')), reverse=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename  = f"scraped_multi_{job['site_type']}_{timestamp}.xlsx"
    filepath  = generate_multi_excel(results_to_export, filename)
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/history')
def get_history():
    jobs = []
    for job in scrape_jobs.values():
        jobs.append({
            'id':           job['id'],
            'url':          job['url'],
            'site_type':    job['site_type'],
            'status':       job['status'],
            'message':      job['message'],
            'result_count': len(job['results']),
            'created_at':   job['created_at'],
            'completed_at': job['completed_at'],
        })
    return jsonify(jobs)


@app.route('/api/database')
def get_database_jobs():
    source  = request.args.get('source') or None
    country = request.args.get('country') or None
    page    = int(request.args.get('page', 1))
    limit   = int(request.args.get('limit', 100))
    total   = database.get_job_count(source, country)
    jobs    = database.get_jobs(source, country, page, limit)
    return jsonify({
        'jobs':     jobs,
        'total':    total,
        'page':     page,
        'limit':    limit,
        'has_more': (page * limit) < total,
    })


@app.route('/api/stats')
def get_stats():
    source = request.args.get('source')
    country = request.args.get('country')
    count = database.get_job_count(source, country)
    
    # Get all jobs for analysis (limited to a reasonable number for speed)
    # For a full dashboard, we might want a specific DB query for stats.
    # But for now, let's just get the count of ICT vs non-ICT.
    conn = database.get_db_connection()
    cursor = conn.cursor()
    where, params = database._build_where(source, country, database.is_postgres())
    
    ict_query = f'SELECT COUNT(*) FROM jobs {where} {"AND" if "WHERE" in where else "WHERE"} is_ict = {"TRUE" if database.is_postgres() else "1"}'
    cursor.execute(ict_query, params)
    ict_count = cursor.fetchone()[0]
    
    cat_query = f'SELECT ict_category, COUNT(*) FROM jobs {where} {"AND" if "WHERE" in where else "WHERE"} is_ict = {"TRUE" if database.is_postgres() else "1"} GROUP BY ict_category'
    cursor.execute(cat_query, params)
    categories = {row[0]: row[1] for row in cursor.fetchall() if row[0]}

    # Count imprecise dates (30+ days)
    imprecise_query = f"SELECT COUNT(*) FROM jobs {where} {'AND' if 'WHERE' in where else 'WHERE'} date_precision = 'imprecise_30_plus'"
    cursor.execute(imprecise_query, params)
    imprecise_count = cursor.fetchone()[0]
    
    conn.close()
    
    return jsonify({
        "total": count,
        "ict_count": ict_count,
        "ict_rate": round((ict_count / count * 100), 2) if count > 0 else 0,
        "imprecise_count": imprecise_count,
        "categories": categories
    })


import re
from datetime import datetime, timedelta

def normalize_posted_date(date_str):
    """
    Normalizes relative dates and returns (normalized_date, precision).
    30+ days ago is flagged as imprecise and returned as None for trend analysis.
    """
    if not date_str:
        return datetime.now().strftime('%Y-%m-%d'), 'precise'
    
    date_str = str(date_str).lower().strip()
    now = datetime.now()
    
    # "30+ days ago" or "1 month ago" -> IMPRECISE
    if '30+' in date_str or 'month' in date_str:
        return None, 'imprecise_30_plus'
    
    # "X days ago"
    day_match = re.search(r'(\d+)\s*day', date_str)
    if day_match:
        return (now - timedelta(days=int(day_match.group(1)))).strftime('%Y-%m-%d'), 'precise'
    
    # "X hours ago" or "X mins ago"
    time_match = re.search(r'(\d+)\s*(hour|min|hr)', date_str)
    if time_match or 'just' in date_str or 'active' in date_str:
        return now.strftime('%Y-%m-%d'), 'precise'
    
    # "X weeks ago"
    week_match = re.search(r'(\d+)\s*week', date_str)
    if week_match:
        return (now - timedelta(weeks=int(week_match.group(1)))).strftime('%Y-%m-%d'), 'precise'

    # If it's already a YYYY-MM-DD string, preserve it
    try:
        if len(date_str) == 10 and date_str[4] == '-' and date_str[7] == '-':
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str, 'precise'
    except:
        pass

    return now.strftime('%Y-%m-%d'), 'precise'


@app.route('/api/analysis')
def get_analysis():
    source = request.args.get('source')
    country = request.args.get('country')
    period = request.args.get('period', 'week') # day, week, month
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = database.get_db_connection()
    cursor = conn.cursor()
    where, params = database._build_where(source, country, database.is_postgres())

    # Add Date range to where clause
    if start_date:
        prefix = " AND " if "WHERE" in where else " WHERE "
        where += f"{prefix}posted >= %s" if database.is_postgres() else f"{prefix}posted >= ?"
        params.append(start_date)
    if end_date:
        prefix = " AND " if "WHERE" in where else " WHERE "
        where += f"{prefix}posted <= %s" if database.is_postgres() else f"{prefix}posted <= ?"
        params.append(end_date)
    
    # Category distribution
    cat_query = f'SELECT ict_category, COUNT(*) as count FROM jobs {where} {"AND" if "WHERE" in where else "WHERE"} is_ict = {"TRUE" if database.is_postgres() else "1"} GROUP BY ict_category ORDER BY count DESC'
    cursor.execute(cat_query, params)
    category_data = [{"category": row[0], "count": row[1]} for row in cursor.fetchall() if row[0]]
    
    # Top employers for ICT
    emp_query = f'SELECT company, COUNT(*) as count FROM jobs {where} {"AND" if "WHERE" in where else "WHERE"} is_ict = {"TRUE" if database.is_postgres() else "1"} GROUP BY company ORDER BY count DESC LIMIT 10'
    cursor.execute(emp_query, params)
    employer_data = [{"company": row[0], "count": row[1]} for row in cursor.fetchall()]
    
    # Portal distribution
    portal_query = f'SELECT source, COUNT(*) as count FROM jobs {where} GROUP BY source'
    cursor.execute(portal_query, params)
    portal_data = [{"portal": row[0], "count": row[1]} for row in cursor.fetchall()]
    
    # ICT Rate by Source
    rate_query = f'SELECT source, COUNT(*) as total, SUM(CASE WHEN is_ict = {"TRUE" if database.is_postgres() else "1"} THEN 1 ELSE 0 END) as ict FROM jobs {where} GROUP BY source'
    try:
        cursor.execute("ALTER TABLE jobs ADD COLUMN is_ict BOOLEAN DEFAULT 0")
        cursor.execute("ALTER TABLE jobs ADD COLUMN ict_category TEXT")
        cursor.execute("ALTER TABLE jobs ADD COLUMN date_precision TEXT DEFAULT 'precise'")
        conn.commit()
    except:
        pass 
    cursor.execute(rate_query, params)
    source_rates = []
    for row in cursor.fetchall():
        total = row[1]
        ict = row[2]
        source_rates.append({
            "source": row[0],
            "total": total,
            "ict": ict,
            "rate": round((ict / total * 100), 2) if total > 0 else 0
        })

    # Country distribution for ICT (normalized — no Uae/UAE split)
    if database.is_postgres():
        country_expr = """CASE LOWER(country)
            WHEN 'uae' THEN 'UAE' WHEN 'ksa' THEN 'KSA'
            WHEN 'qatar' THEN 'Qatar' WHEN 'kuwait' THEN 'Kuwait'
            ELSE COALESCE(NULLIF(country, ''), 'Unknown') END"""
    else:
        country_expr = """CASE LOWER(country)
            WHEN 'uae' THEN 'UAE' WHEN 'ksa' THEN 'KSA'
            WHEN 'qatar' THEN 'Qatar' WHEN 'kuwait' THEN 'Kuwait'
            ELSE CASE WHEN country IS NULL OR country = '' THEN 'Unknown' ELSE country END END"""
    country_query = f'SELECT {country_expr} as country, COUNT(*) as count FROM jobs {where} {"AND" if "WHERE" in where else "WHERE"} is_ict = {"TRUE" if database.is_postgres() else "1"} GROUP BY {country_expr} ORDER BY count DESC'
    cursor.execute(country_query, params)
    country_data = [{"country": row[0] or "Unknown", "count": row[1]} for row in cursor.fetchall()]

    # Weekly/Daily/Monthly trend
    if database.is_postgres():
        trunc_map = {'day': 'day', 'week': 'week', 'month': 'month'}
        trunc = trunc_map.get(period, 'week')
        trend_query = f'''
            SELECT date_trunc('{trunc}', posted::date) as time_unit, COUNT(*) 
            FROM jobs 
            {where} {"AND" if "WHERE" in where else "WHERE"} is_ict = TRUE 
            AND posted IS NOT NULL AND posted != ''
            GROUP BY time_unit ORDER BY time_unit ASC
        '''
    else:
        # SQLite period grouping
        if period == 'day':
            group_fmt = "date(posted)"
        elif period == 'month':
            group_fmt = "strftime('%Y-%m', posted)"
        else: # week
            group_fmt = "date(posted, 'weekday 0', '-6 days')"
            
        trend_query = f'''
            SELECT {group_fmt} as time_unit, COUNT(*) 
            FROM jobs 
            {where} {"AND" if "WHERE" in where else "WHERE"} is_ict = 1 
            AND posted IS NOT NULL AND posted != ''
            AND date_precision = 'precise'
            GROUP BY time_unit ORDER BY time_unit ASC
        '''
    
    try:
        cursor.execute(trend_query, params)
        trend_data = [{"label": row[0], "count": row[1]} for row in cursor.fetchall()]
    except Exception as e:
        print(f"Trend Error: {e}")
        trend_data = []

    conn.close()
    
    return jsonify({
        "category_distribution": category_data,
        "top_employers": employer_data,
        "portal_distribution": portal_data,
        "source_rates": source_rates,
        "country_distribution": country_data,
        "weekly_trend": trend_data
    })


@app.route('/api/database/export')
def export_database():
    source  = request.args.get('source') or None
    country = request.args.get('country') or None
    jobs    = database.get_all_jobs(source, country)
    if not jobs:
        return jsonify({'error': 'No data in database to export'}), 400

    timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
    parts = [p for p in [source, country] if p]
    label = ('_'.join(parts)).lower() if parts else 'all'
    filename   = f"scraped_{label}_db_{timestamp}.xlsx"
    filepath   = generate_excel(jobs, filename)
    return send_file(filepath, as_attachment=True, download_name=filename)

@app.route('/api/database/export_multi')
def export_multi_database():
    source  = request.args.get('source') or None
    country = request.args.get('country') or None
    jobs    = database.get_all_jobs(source, country)
    if not jobs:
        return jsonify({'error': 'No data in database to export'}), 400

    timestamp  = datetime.now().strftime('%Y%m%d_%H%M%S')
    parts = [p for p in [source, country] if p]
    label = ('_'.join(parts)).lower() if parts else 'all'
    filename   = f"scraped_{label}_db_multi_{timestamp}.xlsx"
    filepath   = generate_multi_excel(jobs, filename)
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/api/database/clear', methods=['DELETE'])
def clear_database():
    source  = request.args.get('source') or None
    country = request.args.get('country') or None
    success = database.clear_jobs(source, country)
    if success:
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to clear database'}), 500


@app.route('/api/database/import', methods=['POST'])
def import_database():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    default_source  = request.form.get('default_source', 'Imported')
    default_country = request.form.get('default_country', '')

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        
        total_imported = 0
        total_found = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Infer country/portal from sheet name
            sheet_parts = sheet_name.replace(' ', '_').split('_')
            inferred_country = default_country
            inferred_source = default_source
            
            known_countries = ['UAE', 'KSA', 'Qatar', 'Kuwait']
            known_portals = ['LinkedIn', 'Naukrigulf', 'GulfTalent', 'Bayt']

            for p in sheet_parts:
                if p.upper() in [c.upper() for c in known_countries]:
                    inferred_country = database.normalize_country(p)
                if p.lower() in [s.lower() for s in known_portals]:
                    inferred_source = p

            # Map headers
            headers = [str(cell.value).lower().replace(' ', '_') if cell.value else None for cell in ws[1]]
            
            jobs_to_import = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                
                job = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        job[headers[i]] = str(val) if val is not None else ""
                
                # Normalize keys with more aliases
                title = job.get('title') or job.get('job_title') or job.get('job_name') or job.get('position') or ""
                company = job.get('company') or job.get('company_name') or job.get('employer') or ""
                if not title or not company: continue

                # ICT Classification
                is_ict, ict_cat = classifier.classify_job(title)
                
                # Date Normalization (anchor to scraped_at when available)
                from import_portal_data import resolve_posted_date
                posted_str = job.get('posted') or job.get('posted_date') or job.get('date') or job.get('date_posted') or ""
                scraped_at = job.get('scraped_at') or ""
                norm_date, precision = resolve_posted_date(posted_str, scraped_at)

                job_data = {
                    'title': title,
                    'company': company,
                    'location': job.get('location') or "",
                    'experience': job.get('experience') or "",
                    'posted': norm_date,
                    'source': job.get('source') or inferred_source,
                    'country': database.normalize_country(job.get('country') or inferred_country),
                    'is_ict': is_ict,
                    'ict_category': ict_cat,
                    'date_precision': precision
                }
                jobs_to_import.append(job_data)

            if jobs_to_import:
                inserted = database.save_jobs_batch(jobs_to_import)
                total_imported += inserted
                total_found += len(jobs_to_import)

        return jsonify({'success': True, 'inserted': total_imported, 'total': total_found})

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Import failed: {str(e)}'}), 500


@app.route('/api/delete/<job_id>', methods=['DELETE'])
def delete_job(job_id):
    if job_id in scrape_jobs:
        del scrape_jobs[job_id]
        return jsonify({'success': True})
    return jsonify({'error': 'Job not found'}), 404


@app.route('/api/scrape/delete_results/<job_id>', methods=['DELETE'])
def delete_scrape_results(job_id):
    """Delete all database rows created by a specific scrape job."""
    success = database.delete_jobs_by_scrape_id(job_id)
    if success:
        # Also update the in-memory job to reflect results are gone
        if job_id in scrape_jobs:
            scrape_jobs[job_id]['results'] = []
            scrape_jobs[job_id]['message'] = 'Results deleted from database'
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to delete results'}), 500


if __name__ == '__main__':
    print("\n[*] Scrapling UI is running!")
    print(f"    Concurrent pages : {CONCURRENT_PAGES}")
    print(f"    Stop after dupe pages: {STOP_AFTER_DUPE_PAGES}")
    print("    Open http://localhost:5000 in your browser\n")
    app.run(debug=True, port=5000)