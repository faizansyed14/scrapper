"""
import_portal_data.py — Bulk-import old portal CSV/XLSX exports into scraped_jobs.db.

Resolves relative posted dates (e.g. "2 days ago") against each row's scraped_at
timestamp instead of today's date.

Usage:
    python import_portal_data.py                     # import portal_old_data/
    python import_portal_data.py --dry-run           # preview only
    python import_portal_data.py --folder path/to/dir
"""

import argparse
import csv
import os
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))

import classifier
import database
from database import normalize_country

# ─── Filename → source / country ─────────────────────────────────────────────

CITY_COUNTRY = {
    'abudhabi': 'UAE',
    'dubai': 'UAE',
    'riyadh': 'KSA',
    'jeddah': 'KSA',
    'dammam': 'KSA',
}

PORTAL_NAMES = {
    'bayt': 'Bayt',
    'gulftalent': 'GulfTalent',
    'naukrigulf': 'Naukrigulf',
}

DATE_FORMATS = [
    '%d %b %Y',   # 3 Apr 2026
    '%d %B %Y',   # 3 April 2026
    '%d %b',      # 2 Apr  (year from scraped_at)
    '%d %B',
    '%Y-%m-%d',
    '%d/%m/%Y',
    '%m/%d/%Y',
]


def parse_filename(filename: str) -> tuple[str, str]:
    """Return (source, country) from e.g. abudhabi-bayt2.csv."""
    stem = Path(filename).stem.lower()
    m = re.match(r'^([a-z]+)-(.+)$', stem)
    if not m:
        return 'Imported', ''

    city_key, portal_key = m.group(1), m.group(2)
    portal_key = re.sub(r'\d+$', '', portal_key)
    portal_key = portal_key.replace('naurkigulf', 'naukrigulf')

    source = PORTAL_NAMES.get(portal_key, portal_key.title())
    country = CITY_COUNTRY.get(city_key, '')
    return source, country


def parse_scraped_at(value) -> datetime:
    if not value:
        return datetime.now()
    s = str(value).strip().split('.')[0]
    for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s[:19] if 'T' in fmt or ' ' in fmt else s[:10], fmt)
        except ValueError:
            continue
    return datetime.now()


def resolve_posted_date(posted_raw: str, scraped_at) -> tuple[str | None, str]:
    """
    Convert posted_date using scraped_at as anchor.
    Example: scraped_at=2026-04-03, posted="2 days ago" → 2026-04-01
    """
    anchor = parse_scraped_at(scraped_at)

    if not posted_raw or not str(posted_raw).strip():
        return anchor.strftime('%Y-%m-%d'), 'precise'

    s = str(posted_raw).lower().strip()

    if '30+' in s or re.search(r'\bmonth', s):
        return None, 'imprecise_30_plus'

    if any(x in s for x in ('active', 'just now', 'today')):
        return anchor.strftime('%Y-%m-%d'), 'precise'

    if 'yesterday' in s:
        return (anchor - timedelta(days=1)).strftime('%Y-%m-%d'), 'precise'

    rel = re.search(
        r'(\d+)\s*(min|mins|minute|minutes|hour|hours|hrs|hr|day|days|week|weeks)\s*ago',
        s,
    )
    if rel:
        val = int(rel.group(1))
        unit = rel.group(2)
        if unit.startswith('min'):
            dt = anchor - timedelta(minutes=val)
        elif unit.startswith('hour') or unit.startswith('hr'):
            dt = anchor - timedelta(hours=val)
        elif unit.startswith('day'):
            dt = anchor - timedelta(days=val)
        else:
            dt = anchor - timedelta(weeks=val)
        return dt.strftime('%Y-%m-%d'), 'precise'

    # Bare "X hrs ago" variants without "ago" word in some exports
    rel2 = re.search(r'(\d+)\s*(hrs?|hours?)\b', s)
    if rel2 and 'ago' not in s:
        dt = anchor - timedelta(hours=int(rel2.group(1)))
        return dt.strftime('%Y-%m-%d'), 'precise'

    # Absolute dates: "3 Apr 2026", "1 Apr", "30 Dec" (yearless → rollback if future)
    raw = str(posted_raw).strip()
    for fmt in DATE_FORMATS:
        try:
            if '%Y' not in fmt:
                dt = datetime.strptime(f'{raw} {anchor.year}', f'{fmt} %Y')
                if dt.date() > anchor.date():
                    dt = dt.replace(year=anchor.year - 1)
            else:
                dt = datetime.strptime(raw, fmt)
            return dt.strftime('%Y-%m-%d'), 'precise'
        except ValueError:
            continue

    try:
        if len(s) == 10 and s[4] == '-' and s[7] == '-':
            datetime.strptime(s, '%Y-%m-%d')
            return s, 'precise'
    except ValueError:
        pass

    return anchor.strftime('%Y-%m-%d'), 'precise'


def normalize_header(name: str) -> str:
    return re.sub(r'\s+', '_', (name or '').strip().lower())


def read_csv(path: Path) -> list[dict]:
    rows = []
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({normalize_header(k): (v or '').strip() for k, v in row.items() if k})
    return rows


def read_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [normalize_header(str(c.value) if c.value else '') for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append({
            headers[i]: (str(v).strip() if v is not None else '')
            for i, v in enumerate(row) if i < len(headers) and headers[i]
        })
    return rows


def row_to_job(row: dict, source: str, country: str, scrape_id: str) -> dict | None:
    title = row.get('title') or row.get('job_title') or ''
    company = row.get('company') or row.get('company_name') or ''
    if not title or not company:
        return None

    posted_raw = (
        row.get('posted_date') or row.get('posted') or
        row.get('date') or row.get('date_posted') or ''
    )
    scraped_at = row.get('scraped_at') or ''
    posted, precision = resolve_posted_date(posted_raw, scraped_at)

    is_ict, ict_cat = classifier.classify_job(title)

    return {
        'title': title,
        'company': company,
        'location': row.get('location') or '',
        'experience': row.get('experience') or '',
        'posted': posted,
        'source': source,
        'country': normalize_country(country),
        'scrape_id': scrape_id,
        'is_ict': is_ict,
        'ict_category': ict_cat,
        'date_precision': precision,
        'scraped_at': scraped_at or None,
    }


def import_file(path: Path, dry_run: bool = False) -> dict:
    source, country = parse_filename(path.name)
    scrape_id = f'portal_old_{path.stem.lower()}'

    if path.suffix.lower() == '.csv':
        raw_rows = read_csv(path)
    elif path.suffix.lower() in ('.xlsx', '.xlsm'):
        raw_rows = read_xlsx(path)
    else:
        return {'file': path.name, 'skipped': True, 'reason': 'unsupported format'}

    jobs = []
    for row in raw_rows:
        job = row_to_job(row, source, country, scrape_id)
        if job:
            jobs.append(job)

    inserted = 0
    if jobs and not dry_run:
        inserted = database.save_jobs_batch(jobs, ignore_duplicates=True)

    relative = sum(
        1 for r in raw_rows
        if re.search(r'\bago\b|hrs?\b|hours?\b|days?\b', (r.get('posted_date') or r.get('posted') or ''), re.I)
    )

    return {
        'file': path.name,
        'source': source,
        'country': country,
        'rows': len(raw_rows),
        'valid': len(jobs),
        'inserted': inserted,
        'relative_dates': relative,
    }


def main():
    parser = argparse.ArgumentParser(description='Import old portal CSV/XLSX exports into DB')
    parser.add_argument(
        '--folder',
        default=os.path.join(os.path.dirname(__file__), '..', 'portal_old_data'),
        help='Folder with export files (default: portal_old_data/)',
    )
    parser.add_argument('--dry-run', action='store_true', help='Parse only, no DB writes')
    parser.add_argument(
        '--files',
        nargs='+',
        help='Import only these filenames (in --folder) or full paths',
    )
    args = parser.parse_args()

    folder = Path(args.folder).resolve()
    if not folder.is_dir():
        print(f'[!] Folder not found: {folder}')
        sys.exit(1)

    database.init_db()
    before = database.get_job_count()

    if args.files:
        files = []
        for name in args.files:
            p = Path(name)
            files.append(p if p.is_file() else folder / name)
        files = sorted(set(files))
        missing = [p for p in files if not p.is_file()]
        if missing:
            print('[!] Files not found:')
            for p in missing:
                print(f'    {p}')
            sys.exit(1)
    else:
        files = sorted(
            p for p in folder.iterdir()
            if p.suffix.lower() in ('.csv', '.xlsx', '.xlsm')
        )
    if not files:
        print(f'[!] No CSV/XLSX files in {folder}')
        sys.exit(1)

    print(f'[*] DB before: {before} jobs')
    print(f'[*] Found {len(files)} files in {folder}')
    if args.dry_run:
        print('[*] DRY RUN — no writes\n')

    total_rows = total_valid = total_inserted = total_relative = 0

    for path in files:
        stats = import_file(path, dry_run=args.dry_run)
        if stats.get('skipped'):
            print(f"  SKIP {stats['file']}: {stats['reason']}")
            continue

        total_rows += stats['rows']
        total_valid += stats['valid']
        total_inserted += stats['inserted']
        total_relative += stats['relative_dates']

        ins = stats['inserted'] if not args.dry_run else '(dry-run)'
        print(
            f"  {stats['file']}: {stats['valid']}/{stats['rows']} valid, "
            f"{stats['relative_dates']} relative dates, inserted={ins} "
            f"[{stats['source']}/{stats['country']}]"
        )

    after = database.get_job_count() if not args.dry_run else before
    print(f'\n[+] Parsed {total_valid} jobs from {total_rows} rows ({total_relative} relative dates)')
    if not args.dry_run:
        print(f'[+] Inserted {total_inserted} new jobs (duplicates skipped)')
        print(f'[+] DB after: {after} jobs (+{after - before})')
    else:
        print('[+] Re-run without --dry-run to import')


if __name__ == '__main__':
    main()
